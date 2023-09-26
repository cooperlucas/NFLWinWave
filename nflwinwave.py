import argparse
from base64 import urlsafe_b64encode
from datetime import datetime, timedelta, timezone
from dateutil import parser, tz
from email.mime.text import MIMEText
import espn_scraper as espn
#from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import json
import numpy as np
import os
import pickle
import requests
from requests.adapters import HTTPAdapter, Retry
import scipy.stats as stats
from subprocess import Popen, PIPE
import sys
import threading
import time
import tweepy
from twilio.rest import Client
import traceback
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
import seaborn as sns
from scipy.stats import logistic
import nfl_data_py as nfl
import os
import re
import openpyxl
from matplotlib.ticker import FuncFormatter

# A dictionary of plays that have already been tweeted.
tweeted_plays = None
tweeted_scorecards = None

# A dictionary of the currently active games.
games = {}

# The authenticated Tweepy APIs.
win_wave_api = None
win_wave_client = None

# NPArray of historical win wave moment indices.
historical_win_wave_data = None

# Whether the bot should tweet out any plays
should_tweet = True

### HISTORY FUNCTIONS ###
def has_been_tweeted(play_id, game_id):
    global tweeted_plays
    game_plays = tweeted_plays.get(game_id, [])
    return play_id in game_plays 


def has_been_seen(play_id, game_id):
    global seen_plays
    game_plays = seen_plays.get(game_id, [])
    if play_id in game_plays:
        return True
    game_plays.append(play_id)
    seen_plays[game_id] = game_plays
    return False


def has_been_final(game_id):
    global final_games
    if game_id in final_games:
        return True
    final_games.add(game_id)
    return False


def load_tweeted_plays_dict():
    global tweeted_plays
    tweeted_plays = {}
    if os.path.exists('tweeted_plays.json'):
        file_mod_time = os.path.getmtime('tweeted_plays.json')
    else:
        file_mod_time = 0.
    if time.time() - file_mod_time < 60 * 60 * 12:
        # if file modified within past 12 hours
        with open('tweeted_plays.json', 'r') as f:
            tweeted_plays = json.load(f)
    else:
        with open('tweeted_plays.json', 'w') as f:
            json.dump(tweeted_plays, f)

def load_tweeted_scorecards_dict():
    global tweeted_scorecards
    tweeted_scorecards = {}
    if os.path.exists('tweeted_scorecards.json'):
        file_mod_time = os.path.getmtime('tweeted_scorecards.json')
    else:
        file_mod_time = 0.
    if time.time() - file_mod_time < 60 * 60 * 12:
        # if file modified within past 12 hours
        with open('tweeted_scorecards.json', 'r') as f:
            tweeted_scorecards = json.load(f)
    else:
        with open('tweeted_scorecards.json', 'w') as f:
            json.dump(tweeted_scorecards, f)

def update_tweeted_plays(play_id, game_id):
    global tweeted_plays
    game_plays = tweeted_plays.get(game_id, [])
    game_plays.append(play_id)
    tweeted_plays[game_id] = game_plays
    with open('tweeted_plays.json', 'w') as f:
        json.dump(tweeted_plays, f)
        
def update_tweeted_scorecards(game_id):
    global tweeted_scorecards
    tweeted_scorecards[game_id] = True
    with open('tweeted_scorecards.json', 'w') as f:
        json.dump(tweeted_scorecards, f)
    print('updated tweeted play dict')

### PERCENTILE FUNCTIONS ###

def load_historical_win_wave_data():
    historical_win_wave_data = []

    workbook = openpyxl.load_workbook('historical_game_win_waves.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip the first row (header)
        historical_win_wave_data.append(row)
        
    return historical_win_wave_data

def write_historical_win_wave_data(win_wave_data):
    # Load the Excel workbook or create a new one if it doesn't exist
    workbook = openpyxl.load_workbook('historical_game_win_waves.xlsx')

    # Select the active sheet (usually the first sheet)
    sheet = workbook.active

    # Append the data to the Excel sheet
    sheet.append(win_wave_data)

    # Save the workbook
    workbook.save('historical_game_win_waves.xlsx')
    
def calculate_percentile(game_id, loser_max_wp, num_win_waves, should_update_file=True):
    global historical_win_wave_data
    
    if historical_win_wave_data is not None:
        # Convert the data to a list of tuples for sorting
        data_tuples = [(float(row[1]), int(row[2])) for row in historical_win_wave_data]

        # Extract loser_max_wp and num_win_waves into separate arrays
        loser_max_wp_values = [data[0] for data in data_tuples]
        num_win_waves_values = [data[1] for data in data_tuples]

        # Calculate percentiles
        loser_max_wp_percentile = int(len([value for value in loser_max_wp_values if value < loser_max_wp]) / len(loser_max_wp_values) * 100)
        num_win_waves_percentile = int(len([value for value in num_win_waves_values if value < num_win_waves]) / len(num_win_waves_values) * 100)
              
              
    if should_update_file:
        win_wave_data = [game_id, str(round(loser_max_wp,1)), str(num_win_waves)]
        write_historical_win_wave_data(win_wave_data)

    
    return loser_max_wp_percentile, num_win_waves_percentile

### STRING FORMAT FUNCTIONS ###

def get_qtr_str(qtr):
    if qtr <= 4:
        return 'the ' + str(qtr) + get_ordinal_suffix(qtr)
    elif qtr == 5:
        return 'OT'
    elif qtr == 6:
        return '2 OT'
    elif qtr == 7:
        return '3 OT'
    return ''

def determine_loser(home_score, away_score, home_team_name, away_team_name, home_max_wp, away_max_wp):
    if home_score > away_score:
        return away_team_name, away_max_wp
    elif away_score > home_score:
        return home_team_name, home_max_wp
    else:
        return "Tie", None

def get_ordinal_suffix(num):
    if 10 <= int(num) % 100 <= 20:
        suffix = 'th'
    else:
        last_digit = int(num) % 10
        if last_digit == 1:
            suffix = 'st'
        elif last_digit == 2:
            suffix = 'nd'
        elif last_digit == 3:
            suffix = 'rd'
        else:
            suffix = 'th'
    return suffix

def pretty_play_str(text):
    # Define a regular expression pattern to match parentheses and time format
    pattern_1 = r'\([^)]*\)|\d+:\d+\s'
    
    # Use re.sub to replace all matches of the pattern with an empty string
    cleaned_text = re.sub(pattern_1, '', text)
    
    pattern_2 = r'\[[^\]]*\]'
    
    # Use re.sub to replace all matches of the pattern with an empty string
    cleaned_text = re.sub(pattern_2, '', cleaned_text)
    
    pattern_3 = r'(\w)\s*\.'
    
    cleaned_text = re.sub(pattern_3, r'\1.', cleaned_text)
    
    
    pattern_4 = r'\s*,\s*'
    
    cleaned_text = re.sub(pattern_4, ',', cleaned_text)
    
    # Remove trailing spaces and period
    cleaned_text = cleaned_text.rstrip(' .')
    
    if "TOUCHDOWN" in cleaned_text:
        cleaned_text = cleaned_text.split("TOUCHDOWN")[0] + "TOUCHDOWN!"
    
    if "GOOD" in cleaned_text:
        cleaned_text = cleaned_text.split("GOOD")[0] + "GOOD!"
    
    if "BLOCKED" in cleaned_text:
        cleaned_text = cleaned_text.split("BLOCKED")[0] + "BLOCKED!"
    
    return cleaned_text

def pretty_down_distance_str(down, ydstogo):
    return str(down) + get_ordinal_suffix(down) + ' & ' + str(ydstogo)

### TWITTER FUNCTIONS ###

def initialize_api():
    win_wave_client = tweepy.Client(
        consumer_key="*************",
        consumer_secret="***************************",
        access_token="*************************",
        access_token_secret="*****************************"
    )
    return win_wave_client 

def create_tweet_str(play_desc, 
                     home_score, 
                     away_score, 
                     down,
                     ydstogo,
                     qtr, 
                     qtr_min_remaining,
                     home_wp,
                     home_wp_delta,
                     home_win_wave,
                     away_win_wave,
                     home_team_name,
                     away_team_name,
                     home_hashtag,
                     away_hashtag,
                    ):

    if home_win_wave == 1:
        wavey_str = "WIN WAVE MOMENT" 
        intro_str = "With " + str(qtr_min_remaining) + " left in " + str(get_qtr_str(qtr)) + ", the " + str(home_team_name) + " had a #WinWaveMoment"
        down_str = str(pretty_down_distance_str(down, ydstogo)) + " "
        play_desc_str = str(pretty_play_str(play_desc))[:85]
        delta_str = "This play increased the " + str(home_team_name) + " probability of winning by " + str(round(abs(home_wp_delta*100),1)) + "%, to " + str(round((home_wp * 100),1)) + "%."
        outro_str = str(away_team_name) + " " + str(away_score) + " - " + str(home_team_name) + " " + str(home_score) 
        hashtag_str = str(home_hashtag) + " | " + "#SurfsUp"

        
    elif away_win_wave == 1:
        wavey_str = "WIN WAVE MOMENT" 
        intro_str = "With " + str(qtr_min_remaining) + " left in " + str(get_qtr_str(qtr)) + ", the " + str(away_team_name) + " had a #WinWaveMoment"
        down_str = str(pretty_down_distance_str(down, ydstogo)) + " "
        play_desc_str = str(pretty_play_str(play_desc))[:85]
        delta_str = "This play increased the " + str(away_team_name) + " probability of winning by " + str(round(abs(home_wp_delta*100),1)) + "%, to " + str(round(((1-home_wp) * 100),1)) + "%."
        outro_str = str(away_team_name) + " " + str(away_score) + " - " + str(home_team_name) + " " + str(home_score) 
        hashtag_str = str(away_hashtag) + " | " + "#SurfsUp"
        
    return intro_str + '\n\n' + down_str + play_desc_str + '\n\n' + delta_str + '\n\n' + outro_str + '\n\n' + hashtag_str

def create_scorecard_str(game_id,
                    home_score,
                    away_score, 
                    home_team_name, 
                    away_team_name,
                    home_max_wp,
                    away_max_wp,
                    num_home_win_waves,
                    num_away_win_waves,
                    home_hashtag,
                    away_hashtag):
    
    
    loser, max_win_probability = determine_loser(home_score, away_score, home_team_name, away_team_name, home_max_wp, away_max_wp)
    if loser == "Tie":
        peak_wp_str = "Peak win probabilities: " + str(away_team_name) + " " + str(round(away_max_wp * 100),1) + "% | " + str(home_team_name) + " " + str(round(home_max_wp * 100),1) + "%"
        history_str = ''
    else:
        peak_wp_str = str(loser) + " peak win probability: " + str(round(max_win_probability * 100,1)) + "%"
        loser_max_wp_percentile, num_win_waves_percentile = calculate_percentile(game_id, float(max_win_probability * 100), (num_home_win_waves+num_away_win_waves))
        history_str = 'This comes in at the ' + str(loser_max_wp_percentile) + get_ordinal_suffix(str(loser_max_wp_percentile)) + ' percentile of peak losing team win probability & ' + str(num_win_waves_percentile) + get_ordinal_suffix(str(num_win_waves_percentile)) + " percentile for total number of #WinWaveMoments"

        
    wavey_str = "WIN WAVE SCORECARD" 
    score_str =  str(away_team_name) + " " + str(away_score) + " - " + str(home_team_name) + " " + str(home_score) 
    count_waves_str = "Win Wave moments: " + str(away_team_name) + " " + str(num_away_win_waves) + " | " + str(home_team_name) + " " + str(num_home_win_waves)
    peak_wp_str = peak_wp_str
    history_str = history_str
    hashtag_str = str(away_hashtag) + " | " + str(home_hashtag) + " | " + "#SurfsUp"

    return wavey_str + '\n\n' + score_str + '\n\n' + count_waves_str + '\n' + peak_wp_str + '\n\n' + history_str + '\n\n' + hashtag_str



def tweet_scorecard(game_id,
                    home_score,
                    away_score, 
                    home_team_name, 
                    away_team_name,
                    home_max_wp,
                    away_max_wp,
                    num_home_win_waves,
                    num_away_win_waves,
                    graph_img,
                    home_hashtag,
                    away_hashtag):
    global win_wave_api
    global win_wave_client
    global should_tweet


    scorecard_str = create_scorecard_str(game_id,
                    home_score,
                    away_score, 
                    home_team_name, 
                    away_team_name,
                    home_max_wp,
                    away_max_wp,
                    num_home_win_waves,
                    num_away_win_waves,                      
                    home_hashtag,
                    away_hashtag)
    ##media = win_wave_client.upload_media(media=graph_img)
                                     
    time_print(scorecard_str)
    
    if should_tweet: 
        try:
            update_tweeted_scorecards(game_id)
            tweet = win_wave_client.create_tweet(text=scorecard_str) #, media_ids=[media.media_id])
            print(f"Tweet successfully sent with ID: {tweet['data']['id']}")
           
        except tweepy.TweepyException as e:
            print(f"Tweepy error: {e}")
        except tweepy.errors.HTTPException as e:
            print(f"HTTP request error: {e}")

    


def tweet_play(game_id,
               play_id,
               play_desc, 
               home_score, 
               away_score, 
               down,
               ydstogo,
               qtr, 
               qtr_min_remaining,
               home_wp,
               home_wp_delta,
               home_win_wave,
               away_win_wave,
               home_team_name,
               away_team_name,
               home_hashtag,
               away_hashtag):
    global win_wave_api
    global win_wave_client
    global should_tweet


    tweet_str = create_tweet_str(play_desc, 
               home_score, 
               away_score, 
               down,
               ydstogo,
               qtr, 
               qtr_min_remaining,
               home_wp,
               home_wp_delta,
               home_win_wave,
               away_win_wave,
               home_team_name,
               away_team_name,
               home_hashtag,
               away_hashtag)

    time_print(tweet_str)

    if should_tweet: 
        try:
            update_tweeted_plays(play_id, game_id)
            tweet = win_wave_client.create_tweet(text=tweet_str)
            print(f"Tweet successfully sent with ID: {tweet['data']['id']}")
        except tweepy.TweepyException as e:
            print(f"Tweepy error: {e}")
        except tweepy.errors.HTTPException as e:
            print(f"HTTP request error: {e}")

    def plot_for_data(data, logos, img_path = None):
    # Get game information
    home_team_abbr = logos['home_team'][0]
    away_team_abbr = logos["away_team"][0]

    # Build logo placement data
    logo_placement_data = pd.DataFrame({
        "x": [3575, 3575],
        "y": [0.875, 0.125],
        "team_abbr": [home_team_abbr, away_team_abbr]
    })
    
    logos_df = pd.DataFrame(logos)
    
    # Merge logos
    home_logo_placement_data = logo_placement_data.merge(logos_df, left_on="team_abbr", right_on="home_team")
    away_logo_placement_data = logo_placement_data.merge(logos_df, left_on="team_abbr", right_on="away_team")
    
    # Create the plot
    plt.figure(figsize=(10, 6))
    plt.axhline(y=0.5, color='grey', linewidth=1)
    plt.axvline(x=15*60, color='grey')
    plt.axvline(x=30*60, color='grey')
    plt.axvline(x=45*60, color='grey')
    
    for x, qtr in [(13*60, "Q4"), (28*60, "Q3"), (43*60, "Q2"), (58*60, "Q1")]:
        plt.annotate(qtr, (x, 0.90), textcoords="offset points", xytext=(0,10), ha='center', color='black', fontsize=10)
    
    for idx, row in data[data["away_win_wave"] == 1].iterrows():
        plt.plot(row["game_seconds_remaining"], 0.0, 'k^', markersize=25, color="aqua")

    for idx, row in data[data["home_win_wave"] == 1].iterrows():
        plt.plot(row["game_seconds_remaining"], 1.0, 'kv', markersize=25, color="aqua")

    for idx, row in home_logo_placement_data.iterrows():
        img = plt.imread(row["home_logo"])
        imagebox = OffsetImage(img, zoom=0.08)
        ab = AnnotationBbox(imagebox, (row["x"], row["y"]), frameon=False)
        plt.gca().add_artist(ab)
        
    for idx, row in away_logo_placement_data.iterrows():
        img = plt.imread(row["away_logo"])
        imagebox = OffsetImage(img, zoom=0.08)
        ab = AnnotationBbox(imagebox, (row["x"], row["y"]), frameon=False)
        plt.gca().add_artist(ab)
    
    game_seconds_remaining = data["game_seconds_remaining"].to_numpy()
    home_wp = data["home_wp"].to_numpy()
    for idx, row in data.iterrows():
        plt.plot(game_seconds_remaining, home_wp, color="black")

    plt.xlim(0, 3600)
    plt.ylim(0, 1)
    plt.gca().invert_xaxis()
    plt.xlabel("Quarters")
    plt.ylabel("Home Win Probability")
    #plt.title(f"{game_year} Week {game_week}: {away_team_abbr} at {home_team_abbr}")
    #plt.legend(["WinWave WP", "nflfastR WP"], loc="upper right")
    plt.grid(False)
    plt.xticks([])
    #plt.yticks([0, 0.25, 0.5, 0.75, 1])
    plt.yticks([0, 0.5, 1])
    plt.axhline(y=0, color='grey', linestyle='dotted')
    plt.axhline(y=0.25, color='grey', linestyle='dotted')
    plt.axhline(y=0.75, color='grey', linestyle='dotted')
    plt.axhline(y=1, color='grey', linestyle='dotted')
    plt.gca().yaxis.set_major_formatter(FuncFormatter(percent_formatter))
    
    if img_path:
        # Save the plot as an image
        plt.savefig(img_path, bbox_inches='tight')
    
    if not img_path:
        print("No save path")
    
    # Return the saved image path if it was saved, otherwise return None
    return img_path

### HELPER FUNCTIONS


def calculate_game_seconds(row, qtr):
    remaining_minutes, remaining_seconds = map(int, row.split(':'))
    if qtr in {1, 2, 3, 4}:
        return ((4 - qtr) * 15 * 60) + (remaining_minutes * 60) + remaining_seconds
    else:
        return 0  # OT


def percent_formatter(x, pos):
    return f'{x:.0%}'


def is_win_wave(home_wp_delta, home_wp, threshold=0.1):
    if home_wp_delta >= threshold and (home_wp >= .125 and (home_wp <= .875 or home_wp == 1)):
        return "home_win_wave"
    elif home_wp_delta <= -threshold and ((home_wp >= .125 or home_wp == 0) and home_wp <= .875):
        return "away_win_wave"
    else:
        return "no_wave"
    
    
def get_team_hashtag(team_name):
    team_hashtag = {
        "Cardinals": "#BirdGang",
        "Falcons": "#DirtyBirds",
        "Ravens": "#RavensFlock",
        "Bills": "#BillsMafia",
        "Panthers": "#KeepPounding",
        "Bears": "#DaBears",
        "Bengals": "#RuleTheJungle",
        "Browns": "#DawgPound",
        "Cowboys": "#DallasCowboys",
        "Broncos": "#BroncosCountry",
        "Lions": "#OnePride",
        "Packers": "#GoPackGo",
        "Texans": "#WeAreTexans",
        "Colts": "#ForTheShoe",
        "Jaguars": "#DUUUVAL",
        "Chiefs": "#ChiefsKingdom",
        "Raiders": "#RaiderNation",
        "Chargers": "#BoltUp",
        "Rams": "#RamsHouse",
        "Dolphins": "#FinsUp",
        "Vikings": "#Skol",
        "Patriots": "#ForeverNE",
        "Saints": "#Saints",
        "Giants": "#NYGiants",
        "Jets": "#TakeFlight",
        "Eagles": "#FlyEaglesFly",
        "Steelers": "#HereWeGo",
        "49ers": "#FTTB",
        "Seahawks": "#Seahawks",
        "Buccaneers": "#GoBucs",
        "Titans": "#Titans",
        "Commanders": "#HTTC"
    }
    return team_hashtag.get(team_name, "")

    
    
### CURRENT GAME FUNCTIONS ###


def time_print(message):
    print(get_current_time_str() + ": " + str(message))


def get_current_time_str():
    return datetime.now().strftime("%b %-d at %-I:%M:%S %p")


def get_now():
    return datetime.now(tz=tz.gettz())

def game_is_final(play_desc):
    if play_desc == "END GAME":
        return True
    else:
        return False

def update_current_week_games():
    global current_week_games
    current_week_games = []

    espn_data = requests.get(
        "http://site.api.espn.com/apis/site/v2/sports/football/nfl/scoreboard",
        timeout=10).json()
    for event in espn_data['events']:
        current_week_games.append(event)


def get_active_game_ids():
    global current_week_games
    global completed_game_ids
 
    now = get_now()
    active_game_ids = set()

    for game in current_week_games:
        if game['id'] in completed_game_ids:
            # ignore any games that are marked completed (which is done by
            # checking if ESPN says final)
            continue
        game_time = parser.parse(
            game['date']).replace(tzinfo=timezone.utc).astimezone(tz=None)
        if game_time - timedelta(minutes=15) < now and game_time + timedelta(
                hours=6) > now:
            # game should start within 15 minutes and not started more than 6 hours ago
            active_game_ids.add(game['id'])

    return active_game_ids


def download_data_for_active_games():
    global games
    active_game_ids = get_active_game_ids()
    if len(active_game_ids) == 0:
        time_print("No games active. Sleeping for 15 minutes...")
        time.sleep(14 * 60)  # We sleep for another minute in the live callback
    games = {}
    for game_id in active_game_ids:
        base_link = "https://cdn.espn.com/core/nfl/game?xhr=1&gameId="
        game_link = base_link + game_id
        games[game_id] = requests.get(game_link, timeout=10).json()
        
        
        live_callback()

        
### MAIN FUNCTIONS ###


def live_callback():
    global games
    start_time = time.time()
    win_percentage_data = []
    
    # List of play_type_abbr values to exclude
    excluded_play_types = ['Off TO', 'TO', 'EP', '2Min Warn', 'K']
    
    for game_id, game in games.items():
        win_percentage_data = []
        time_print('Getting data for game ID ' + game_id)
        game_info = game["gamepackageJSON"]['boxscore']['teams']

        home_team_data = game_info[1].get('team', {})
        away_team_data = game_info[0].get('team', {})
            
        logos = []
        team_info = {
        "home_team": home_team_data.get('abbreviation'),
        "away_team": away_team_data.get('abbreviation'),
        "home_team_name": home_team_data.get('name'),
        "away_team_name": away_team_data.get('name'),    
        "home_logo": home_team_data.get('logo'),
        "away_logo": away_team_data.get('logo'),
        "home_color": home_team_data.get('color'),
        "away_color": away_team_data.get('color'),
        }
        
        logos.append(team_info)
        
        previous_play_wp = None
        for play in game["gamepackageJSON"]["winprobability"]:

            play_by_play_data = {
                "qtr": play.get('play', {}).get('period', {}).get('number'),
                "home_score": play.get('play', {}).get('homeScore'),
                "away_score": play.get('play', {}).get('awayScore'),
                "down": play.get('play', {}).get('start', {}).get('down'),
                "ydstogo": play.get('play', {}).get('start', {}).get('distance'),
                "ydsToEndZone": play.get('play', {}).get('start', {}).get('yardsToEndzone'),
                "yardLine": play.get('play', {}).get('start', {}).get('yardLine'),
                "qtr_min_remaining": play.get('play', {}).get('clock', {}).get('displayValue'),
                "play_desc": play.get('play', {}).get('text'),
                "play_type": play.get('play', {}).get('type', {}).get('text'),
                "play_type_abbr": play.get('play', {}).get('type', {}).get('abbreviation'),
                "home_wp": play.get('homeWinPercentage'),
                }
            
            # Add previous play wp and absolute difference
            if previous_play_wp is not None:
                play_by_play_data["previous_play_home_wp"] = previous_play_wp
                play_by_play_data["home_wp_delta"] = (play_by_play_data["home_wp"] - previous_play_wp)

            # Update previous_play_wp for the next iteration
            previous_play_wp = play_by_play_data["home_wp"]
            
            try:
                win_percentage_data.append(play_by_play_data)
            except:
                continue
        # Add all plays from game to DF
        win_percentage_data = pd.DataFrame(win_percentage_data)
        
        # Add game seconds remaining column
        win_percentage_data['game_seconds_remaining'] = win_percentage_data.apply(lambda row: calculate_game_seconds(row['qtr_min_remaining'], row['qtr']), axis=1)

        # Check for Win Wave
        win_percentage_data['win_wave'] = win_percentage_data.apply(lambda row: is_win_wave(row['home_wp_delta'], row['home_wp']), axis=1)
        win_percentage_data['home_win_wave'] = win_percentage_data['win_wave'].apply(lambda x: 1 if x == 'home_win_wave' else 0)
        win_percentage_data['away_win_wave'] = win_percentage_data['win_wave'].apply(lambda x: 1 if x == 'away_win_wave' else 0)

        logos_data = {
        "home_team": [item["home_team"] for item in logos],
        "away_team": [item["away_team"] for item in logos],
        "home_team_name": [item["home_team_name"] for item in logos],
        "away_team_name": [item["away_team_name"] for item in logos],
        "home_logo": [item["home_logo"] for item in logos],
        "away_logo": [item["away_logo"] for item in logos],
        "home_color": [item["home_color"] for item in logos],
        "away_color": [item["away_color"] for item in logos],
        }
        
        file_path = 'test.csv'

        # Use the to_csv() method to save the DataFrame to a CSV file
        win_percentage_data.to_csv(file_path, index=False)

        
        for idx, play in win_percentage_data.iterrows():
            # Create a 'play_id'
            play_id = (game_id + 
                               '_' +
                               str(play['game_seconds_remaining']))

            if play['win_wave'] == 'no_wave':
                if game_is_final(play['play_desc']) is True:
                    if has_been_final(game_id):
                        completed_game_ids.add(game_id)
                    else:
                      # Create the Win Wave Graph
                        graph_img = plot_for_data(win_percentage_data, logos_data, f"plot_{game_id}.png")
                        filtered_win_percentage_data = win_percentage_data[~win_percentage_data['play_type_abbr'].isin(excluded_play_types)]
                        try:
                         # send scorecard tweet
                            tweet_scorecard(game_id,
                               play['home_score'],
                               play['away_score'], 
                               logos_data["home_team_name"][0], 
                               logos_data["away_team_name"][0],
                               max(win_percentage_data['home_wp']),
                               1-min(win_percentage_data['home_wp']),       
                               sum(filtered_win_percentage_data['home_win_wave']),
                               sum(filtered_win_percentage_data['away_win_wave']),    
                               graph_img,
                               get_team_hashtag(logos_data["home_team_name"][0]),
                               get_team_hashtag(logos_data["away_team_name"][0]))
                            completed_game_ids.add(game_id)
                        #break
                        except BaseException as e:
                            traceback.print_exc()
                            time_print("Error occurred 2:")
                            time_print(e)
                        
                else: # Not a tweetable play
                    continue
            if has_been_tweeted(play_id, game_id):
                # Play already tweeted 
                continue
            if has_been_seen(play_id, game_id):
                # Play already seen by algorithm
                continue
            
            else:
                # Handle a play with a win wave
                try:
                    if play['play_type_abbr'] in (excluded_play_types):
                        continue # Don't tweet no play/extra points
                except:
                    continue
                try:
                    tweet_play(game_id,
                               play_id,
                               play['play_desc'], 
                               play['home_score'],
                               play['away_score'], 
                               play['down'],
                               play['ydstogo'],
                               play['qtr'], 
                               play['qtr_min_remaining'],
                               play['home_wp'],
                               play['home_wp_delta'],
                               play['home_win_wave'],
                               play['away_win_wave'], 
                               logos_data["home_team_name"][0], 
                               logos_data["away_team_name"][0],
                               get_team_hashtag(logos_data["home_team_name"][0]),
                               get_team_hashtag(logos_data["away_team_name"][0]))
                except BaseException as e:
                    continue

                
    while (time.time() < start_time + 30):
        time.sleep(1)
    print("")
    
def main():
    global win_wave_api
    global win_wave_client
    global historical_win_wave_data
    global should_text
    global should_tweet
    global notify_using_native_mail
    global notify_using_twilio
    global final_games
    global debug
    global not_headless
    global sleep_time
    global seen_plays
    global gmail_client
    global twilio_client
    global completed_game_ids

    
    win_wave_client = initialize_api()
    historical_win_wave_data = load_historical_win_wave_data()

    sleep_time = 1

    completed_game_ids = set()
    final_games = set()
    
    should_continue = True
    while should_continue:
        try:
            update_current_week_games()
            load_tweeted_plays_dict()
            load_tweeted_scorecards_dict()
            seen_plays = {}

            
            now = datetime.now()
            if now.hour < 5:
                stop_date = now.replace(hour=5, minute=0, second=0, microsecond=0)
            else:
                now += timedelta(days=1)
                stop_date = now.replace(hour=5, minute=0, second=0, microsecond=0)

            while datetime.now() < stop_date:
                # Perform your periodic tasks here
                start_time = time.time()
                download_data_for_active_games()
                sleep_time = 1

        except KeyboardInterrupt:
            should_continue = False
        except Exception as e:
            # When an exception occurs: log it, send a message, and sleep for an
            # exponential backoff time
            traceback.print_exc()
            time_print("Error occurred 1:")
            time_print(e)
            time_print("Sleeping for " + str(sleep_time) + " minutes")
            time.sleep(sleep_time * 60)
            #sleep_time *= 2

if __name__ == "__main__":
    main() 


    

